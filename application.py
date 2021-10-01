from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook


 
app = Flask(__name__)

@app.route('/')
def homepage():
    try:
        excel = load_workbook('task.xlsx')
    except:
        excel = Workbook()
    page = excel.active
    goods = [cell.value for cell in page["A"]]
    excel.save('task,xlsx')
        
    return render_template('index.html', tovar = goods)


@app.route('/add/', methods=["POST"])
def add():
    good = request.form['good']
    try:
        excel = excel = load_workbook('task.xlsx')
    except:
        excel = Workbook()

    page = excel.active
    page.append([good])
    excel.save("task.xlsx")
    
   
        

    excel.save('task.xlsx')
    return "<h1>Инвентарь пополнен</h1> <a href='/'>Домой</a>"



if __name__ == "__main__":
    app.run()


